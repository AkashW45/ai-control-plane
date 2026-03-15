# services/runbook_export.py
# Enterprise Excel export — matches MS template format
# Structure: header info → deployment matrix → group sections → test cases per ticket

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DARK_BLUE  = "1F4E78"
LIGHT_BLUE = "D9E1F2"
RED_LIGHT  = "FDECEA"
GREEN_LIGHT= "E8F5E9"
ORANGE_LIGHT="FFF3E0"
PURPLE_LIGHT="EDE7F6"
WHITE      = "FFFFFF"

GROUP_COLORS = {
    "migration":  ("FFF3E0", "B7860B"),
    "bugfix":     ("FDECEA", "C0392B"),
    "feature":    ("E8F5E9", "1A6B3A"),
    "deployment": ("D9E1F2", "1F4E78"),
    "testing":    ("EDE7F6", "4A235A"),
}

def _bold(ws, row, col, value, fill=None, size=11, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=size, bold=True, color=color)
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c

def _val(ws, row, col, value, wrap=True, size=11, color="000000", bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    c.font = Font(name="Calibri", size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    return c

def _thdr(ws, row, col, value, bg=DARK_BLUE):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def _section(ws, row, title, ncols=9, bg=LIGHT_BLUE, text_color="000000"):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", size=11, bold=True, color=text_color)
    c.fill = PatternFill("solid", start_color=bg)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 20
    return row + 1

def _kv(ws, row, key, value, key_bold=True):
    _bold(ws, row, 1, key) if key_bold else _val(ws, row, 1, key)
    _val(ws, row, 2, value)
    return row + 1

def _numbered_list(ws, row, items):
    for i, item in enumerate(items, 1):
        _val(ws, row, 1, f"{i}.", bold=True)
        c = ws.cell(row=row, column=2, value=str(item))
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(30, len(str(item))//60 * 15 + 15)
        row += 1
    return row


def export_runbook_excel(ticket, test_cases=None, plan=None,
                          health_checks=None, sprint_summaries=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Release Runbook"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 18

    row = 1

    # ── Header ────────────────────────────────────────────
    row = _kv(ws, row, "Change ID",          ticket.get("key",""))
    row = _kv(ws, row, "Release Description", ticket.get("summary",""))
    row = _kv(ws, row, "Priority",            ticket.get("priority",""))
    row = _kv(ws, row, "Project",             ticket.get("project",""))
    row = _kv(ws, row, "Fix Version",         ticket.get("fixVersion",""))
    row = _kv(ws, row, "Status",              ticket.get("status",""))
    row = _kv(ws, row, "Generated",           datetime.now().strftime("%Y-%m-%d %H:%M"))
    row += 1

    # ── Deployment Matrix ──────────────────────────────────
    row = _section(ws, row, "Deployment Matrix")
    for col, hdr in enumerate(["Deployment Type","Component","Part of Release","Active Rail",
         "Rail to be Released","Release Version (QA)","Rollback Version (PROD)",
         "Comparison Link","Remarks"], start=1):
        _thdr(ws, row, col, hdr)
    ws.row_dimensions[row].height = 30
    row += 1

    tickets_list = sprint_summaries or [{"key": ticket.get("key",""), "summary": ticket.get("summary",""), "issue_type": ticket.get("issuetype","Task")}]
    for t in tickets_list:
        _val(ws, row, 1, t.get("issue_type","Task"))
        _val(ws, row, 2, t.get("key",""))
        _val(ws, row, 3, "Yes")
        _val(ws, row, 4, "QA")
        _val(ws, row, 5, "PROD")
        _val(ws, row, 6, ticket.get("fixVersion",""))
        _val(ws, row, 7, "")
        _val(ws, row, 8, "")
        _val(ws, row, 9, t.get("summary","")[:80])
        row += 1
    row += 1

    # ── Runbook Summary ────────────────────────────────────
    if plan:
        row = _section(ws, row, "Runbook Summary")
        row = _kv(ws, row, "Summary",        plan.get("summary",""))
        ws.row_dimensions[row-1].height = max(45, len(plan.get("summary",""))//55 * 15 + 15)
        row = _kv(ws, row, "Probable Cause", plan.get("probable_cause",""))
        row += 1

        row = _section(ws, row, "Global Pre-Deployment Checks")
        row = _numbered_list(ws, row, plan.get("pre_checks", []))
        row += 1

        row = _section(ws, row, "Global Resolution Steps")
        for idx, step in enumerate(plan.get("steps",[]), 1):
            _bold(ws, row, 1, f"Step {idx}: {step.get('description','')}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            cmds = "\n".join(step.get("commands",[]))
            c = ws.cell(row=row, column=2, value=cmds)
            c.font = Font(name="Courier New", size=9, color="CDD6F4")
            c.fill = PatternFill("solid", start_color="1E1E2E")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(30, len(step.get("commands",[])) * 14)
            row += 1
        row += 1

    # ── Group Sections ─────────────────────────────────────
    groups = plan.get("groups", []) if plan else []
    if groups:
        row = _section(ws, row, "GROUP-SPECIFIC SECTIONS — Per Ticket Type", bg="2C3E50", text_color=WHITE)
        row += 1

        for grp in groups:
            gtype    = grp.get("type","deployment")
            bg_color, hdr_color = GROUP_COLORS.get(gtype, ("D9E1F2","1F4E78"))
            gname    = grp.get("name","")
            gtickets = ", ".join(grp.get("tickets",[]))

            # Group header
            row = _section(ws, row,
                f"GROUP: {gname.upper()} ({gtype.upper()}) — Tickets: {gtickets}",
                bg=hdr_color, text_color=WHITE)

            # Pre-checks
            if grp.get("pre_checks"):
                row = _section(ws, row, f"  Pre-checks — {gname}", bg=bg_color)
                row = _numbered_list(ws, row, grp["pre_checks"])

            # Steps
            if grp.get("steps"):
                row = _section(ws, row, f"  Steps — {gname}", bg=bg_color)
                for idx, step in enumerate(grp["steps"], 1):
                    _bold(ws, row, 1, f"  Step {idx}: {step.get('description','')}")
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                    row += 1
                    cmds = "\n".join(step.get("commands",[]))
                    c = ws.cell(row=row, column=2, value=cmds)
                    c.font = Font(name="Courier New", size=9, color="CDD6F4")
                    c.fill = PatternFill("solid", start_color="1E1E2E")
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                    ws.row_dimensions[row].height = max(25, len(step.get("commands",[])) * 14)
                    row += 1

            # Validation
            if grp.get("validation"):
                row = _section(ws, row, f"  Validation — {gname}", bg=bg_color)
                row = _numbered_list(ws, row, grp["validation"])

            # Rollback
            if grp.get("rollback"):
                row = _section(ws, row, f"  Rollback — {gname}", bg="FDECEA")
                row = _kv(ws, row, "Rollback Steps", grp["rollback"])
                ws.row_dimensions[row-1].height = max(30, len(grp["rollback"])//60 * 15 + 15)

            row += 1

    # ── Global Validation & Rollback ──────────────────────
    if plan:
        row = _section(ws, row, "Validation & Checks")
        row = _numbered_list(ws, row, plan.get("validation",[]))
        row += 1

        row = _section(ws, row, "Rollback")
        row = _kv(ws, row, "Rollback Strategy", plan.get("rollback",""))
        row = _kv(ws, row, "Rollback SLA",       "< 15 minutes")
        row = _kv(ws, row, "Escalation",         plan.get("escalation",""))
        row += 1

    # ── Test Cases by Ticket ───────────────────────────────
    # Get all test cases from sprint or single ticket
    from_sprint = test_cases and isinstance(test_cases, dict) and "test_suite" in test_cases

    if sprint_summaries:
        # Per-ticket test cases
        sd_sprint_tc = {}
        # Try to get from session (passed via sprint_summaries context)
        all_tests = test_cases.get("test_suite", []) if test_cases else []

        if all_tests:
            row = _section(ws, row, "Test Cases (All Tickets)")
            for col, hdr in enumerate(["ID","Ticket","Category","Priority","Title","Steps","Expected Result"], start=1):
                _thdr(ws, row, col, hdr)
            ws.row_dimensions[row].height = 28
            row += 1
            pri_colors = {"P1":"C0392B","P2":"B7860B","P3":"1A6B3A"}
            for tc in all_tests:
                ticket_key = tc.get("ticket_key", tc.get("id","").split("-TC-")[0] if "-TC-" in tc.get("id","") else "")
                _val(ws, row, 1, tc.get("id",""), bold=True, color="2A7AE4")
                _val(ws, row, 2, ticket_key)
                _val(ws, row, 3, tc.get("category",""))
                _val(ws, row, 4, tc.get("priority",""), bold=True,
                     color=pri_colors.get(tc.get("priority",""),"000000"))
                _val(ws, row, 5, tc.get("title",""))
                steps_txt = "\n".join([f"{i+1}. {s}" for i,s in enumerate(tc.get("steps",[]))])
                _val(ws, row, 6, steps_txt)
                _val(ws, row, 7, tc.get("expected_result",""))
                ws.row_dimensions[row].height = max(30, len(tc.get("steps",[])) * 15)
                row += 1
    elif test_cases and test_cases.get("test_suite"):
        row = _section(ws, row, "Test Cases")
        for col, hdr in enumerate(["ID","Category","Priority","Title","Steps","Expected Result"], start=1):
            _thdr(ws, row, col, hdr)
        ws.row_dimensions[row].height = 28
        row += 1
        pri_colors = {"P1":"C0392B","P2":"B7860B","P3":"1A6B3A"}
        for tc in test_cases["test_suite"]:
            _val(ws, row, 1, tc.get("id",""), bold=True, color="2A7AE4")
            _val(ws, row, 2, tc.get("category",""))
            _val(ws, row, 3, tc.get("priority",""), bold=True,
                 color=pri_colors.get(tc.get("priority",""),"000000"))
            _val(ws, row, 4, tc.get("title",""))
            steps_txt = "\n".join([f"{i+1}. {s}" for i,s in enumerate(tc.get("steps",[]))])
            _val(ws, row, 5, steps_txt)
            _val(ws, row, 6, tc.get("expected_result",""))
            ws.row_dimensions[row].height = max(30, len(tc.get("steps",[])) * 15)
            row += 1
    row += 1

    # ── Health Checks ──────────────────────────────────────
    if health_checks and health_checks.get("health_checks"):
        row = _section(ws, row, "Health Checks (P1/P2 Auto-Generated)")
        for col, hdr in enumerate(["HC ID","Test Case Ref","Title","Command","Expected Output","On Failure"], start=1):
            _thdr(ws, row, col, hdr)
        ws.row_dimensions[row].height = 28
        row += 1
        for hc in health_checks["health_checks"]:
            _val(ws, row, 1, hc.get("id",""), bold=True, color="2A7AE4")
            _val(ws, row, 2, hc.get("test_case_ref",""))
            _val(ws, row, 3, hc.get("title",""))
            c = ws.cell(row=row, column=4, value=hc.get("command",""))
            c.font = Font(name="Courier New", size=9, color="CDD6F4")
            c.fill = PatternFill("solid", start_color="1E1E2E")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            _val(ws, row, 5, hc.get("expected_output",""))
            _val(ws, row, 6, hc.get("on_failure",""), color="C0392B")
            ws.row_dimensions[row].height = 45
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()